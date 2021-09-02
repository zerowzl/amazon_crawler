import time
import urllib

import scrapy
from openpyxl import load_workbook

from ..items import AmazonCrawlerItem


class AmazonSearchSpider(scrapy.Spider):
    name = 'amazonSearch'
    items = []

    def start_requests(self):
        keys = self.getExcelSearchKeys()
        # print(keys)

        urls = [
            'http://www.amazon.com/s'
        ]

        cookies = {"session-id": "145-3683219-9272551"}

        for key in keys:
            if key is not None:
                url_key = key.replace(" ", "+")
                params = {
                    "k": url_key
                }
                url = f'{urls[0]}?{urllib.parse.urlencode(params)}'
                yield scrapy.Request(url=url,
                                     callback=self.parse,
                                     # cookies=cookies,
                                     meta={"key": key})

    def parse(self, response):
        # 打印基础信息
        # print(f"response.status: {response.status}")
        # r = response.request
        # print(r.headers)

        key = response.meta["key"]

        # 是否触发了 Amazon 的反爬机制
        robot_message = response.xpath('/html/body/div/div[1]/div[2]/div/p/text()').get()
        if robot_message is not None:
            # 被认定为机器人, 需要输入验证码
            # print(f"Amazon 反爬 message: {robot_message}")
            # img = response.xpath("/html/body/div/div[1]/div[3]/div/div/form/div[1]/div/div/div[1]/img/@src").extract()
            # print(f"Amazon 反爬 image: {img}")
            # 重试
            yield scrapy.Request(url=response.url, meta={"key": key}, dont_filter=True)

        # 获取搜索结果
        result = response.xpath('//*[@id="search"]/span/div/span/h1/div/div[1]/div/div/span[1]/text()').get()
        print(key)
        print(result)
        if result is None:
            print(f"'{key}'未获取到搜索结果")
            return

        # 1-16 of over 10,000 results for
        # 1-16 of 148 results for
        # 解析数量
        num = result.split(" ")[-3]
        print(f"'{key}'搜索结果数量 : {num}")

        length = len(num)
        # print(length)

        if length < 4:
            print("搜索结果小于1000")
            info = AmazonCrawlerItem()
            info["keyword"] = key
            info["result"] = num
            yield info
        # else:
        #     print("搜索结果大于1000")

    def getExcelSearchKeys(self) -> list:
        # 加载Excel
        start_time = time.time()
        filename = "D:\\file\\搜索词 - 副本.xlsx"
        wb = load_workbook(filename)
        end_time = time.time()
        print(f"完成解析Excel: {end_time - start_time} 秒")

        # 获取第一个 Sheet
        first_sheet = wb.sheetnames[0]
        ws = wb[first_sheet]

        # 解析搜索词
        keys = []
        for row in range(3, ws.max_row + 1):
            for column in "A":
                cell_name = "{}{}".format(column, row)
                keys.append(ws[cell_name].value)
        return keys
